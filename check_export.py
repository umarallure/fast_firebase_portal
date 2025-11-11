import pandas as pd
import openpyxl

def check_export_file():
    try:
        # Read the Excel file
        df = pd.read_excel('test_export.xlsx', sheet_name='Opportunities')
        
        print(f"Excel file loaded successfully!")
        print(f"Shape: {df.shape[0]} rows, {df.shape[1]} columns")
        print(f"\nAll Columns ({len(df.columns)}):")
        for i, col in enumerate(df.columns, 1):
            print(f"{i:2d}. {col}")
        
        # Check if notes columns exist
        notes_columns = ['note1', 'note2', 'note3', 'note4']
        existing_notes_cols = [col for col in notes_columns if col in df.columns]
        print(f"\nNotes columns found: {existing_notes_cols}")
        
        # Also check using openpyxl directly
        print(f"\n--- Checking with openpyxl ---")
        wb = openpyxl.load_workbook('test_export.xlsx')
        ws = wb['Opportunities']
        
        # Get headers from first row
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value is not None:
                headers.append(str(cell_value))
        
        print(f"Openpyxl headers ({len(headers)}): {headers}")
        openpyxl_notes = [h for h in headers if 'note' in h.lower()]
        print(f"Openpyxl notes columns: {openpyxl_notes}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    check_export_file()